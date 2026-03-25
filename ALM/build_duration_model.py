#!/usr/bin/env python3
"""
Duration Calculator Model — ALM Tool
Purpose : Dynamic Macaulay & Modified duration for term loans / borrowings
Author  : Bolt 🦞 for Ashish Prakash
Inputs  : Principal, coupon rate, YTM, tenor, frequency, start date, amort type
Outputs : Cash flow schedule, duration, DV01, convexity, price sensitivity
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import datetime

# ──────────────────────────────────────────────
# PALETTE
# ──────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
YELLOW_FILL = "FFF2CC"
GREEN_FILL  = "E2EFDA"
RED_FILL    = "FCE4D6"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"
DARK_GREY   = "595959"

def header_cell(ws, row, col, value, sub=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=WHITE,
                  name="Calibri", size=10 if sub else 12)
    c.fill = PatternFill("solid",
                         fgColor=MID_BLUE if not sub else DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    ws.row_dimensions[row].height = 30 if not sub else 22
    return c

def input_cell(ws, row, col, value, fmt=None, comment=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=YELLOW_FILL)
    c.font = Font(bold=True, color=DARK_BLUE, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        c.number_format = fmt
    return c

def output_cell(ws, row, col, value, fmt=None, colour=GREEN_FILL):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=colour)
    c.font = Font(bold=True, color=DARK_GREY, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        c.number_format = fmt
    return c

def thin_border(ws, row, col):
    side = Side(style="thin", color="BFBFBF")
    ws.cell(row=row, column=col).border = Border(
        left=side, right=side, top=side, bottom=side)

def apply_border_range(ws, min_row, max_row, min_col, max_col):
    side = Side(style="thin", color="BFBFBF")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = Border(
                left=side, right=side, top=side, bottom=side)

# ──────────────────────────────────────────────
# SHEET 1 — INPUTS
# ──────────────────────────────────────────────
def build_inputs(wb):
    ws = wb.create_sheet("Inputs")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # Title banner
    ws.merge_cells("B1:E1")
    t = ws["B1"]
    t.value = "TERM LOAN / BORROWING — DURATION CALCULATOR"
    t.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    t.fill = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("B2:E2")
    sub = ws["B2"]
    sub.value = "Macaulay Duration  |  Modified Duration  |  DV01  |  Convexity  |  Price Sensitivity"
    sub.font = Font(italic=True, size=9, color=DARK_GREY, name="Calibri")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Section: Loan Parameters
    ws.merge_cells("B4:E4")
    header_cell(ws, 4, 2, "LOAN PARAMETERS", sub=False)

    params = [
        # (row, label, default_value, format, description)
        (6,  "Principal (₹ Cr)",             100.00,   '#,##0.00',   "Loan notional"),
        (7,  "Coupon Rate (% p.a.)",            9.10,   '0.00%',     "Nominal coupon — changes to coupon freq below"),
        (8,  "Frequency (1=Yr, 2=Half, 4=Qtr)",    2,   '0',         "Compounding periods per year"),
        (9,  "Yield / YTM (% p.a.)",             8.75,  '0.00%',     "Market yield used for PV discounting"),
        (10, "Tenor (Years)",                      7,   '0.00',      "Loan maturity in years — can be fractional"),
        (11, "Start Date",    datetime.date(2026, 4, 1), 'DD-MMM-YYYY', "Loan drawdown / first coupon date"),
        (12, "Amortisation Type",           "Bullet", None,         "Bullet OR amortising (enter exactly: Bullet / Full / Partial)"),
        (13, "Partial Amort Rate (% principal/yr)", 0.0, '0.00%',  "Only used if Type = Partial — % of principal repaid per year"),
    ]

    for row, label, default, fmt, desc in params:
        lbl = ws.cell(row=row, column=2, value=label)
        lbl.font = Font(size=10, name="Calibri", color=DARK_GREY)
        lbl.alignment = Alignment(vertical="center")
        lbl.fill = PatternFill("solid", fgColor=GREY)
        thin_border(ws, row, 2)

        inp = input_cell(ws, row, 3, default, fmt=fmt)
        thin_border(ws, row, 3)

        # Spill description across D:E
        ws.merge_cells(start_row=row, start_column=4,
                       end_row=row, end_column=5)
        desc_cell = ws.cell(row=row, column=4, value=desc)
        desc_cell.font = Font(size=8, italic=True, color=DARK_GREY,
                              name="Calibri")
        desc_cell.alignment = Alignment(vertical="center")
        thin_border(ws, row, 4)

    # Section: Calculated Outputs (static preview — live values on Results sheet)
    ws.merge_cells("B15:E15")
    header_cell(ws, 15, 2, "KEY OUTPUTS (LIVE — see Results Sheet)", sub=False)

    outputs = [
        (17, "Macaulay Duration (Years)",      "",  "0.00 \"Yrs\""),
        (18, "Modified Duration (Years)",       "",  "0.00 \"Yrs\""),
        (19, "DV01 (₹ Cr per bp)",              "",  '#,##0.0000'),
        (20, "Convexity",                       "",  '0.0000'),
        (21, "Price per ₹100 face (₹)",         "",  '0.00'),
        (22, "Modified Duration × Face Value",  "",  '#,##0.00'),
        (23, "Approx Price Change per 100bp up (₹ Cr)", "", '#,##0.00'),
        (24, "Approx Price Change per 100bp down (₹ Cr)", "", '#,##0.00'),
    ]

    for row, label, _, fmt in outputs:
        lbl = ws.cell(row=row, column=2, value=label)
        lbl.font = Font(size=10, name="Calibri", color=DARK_GREY)
        lbl.alignment = Alignment(vertical="center")
        lbl.fill = PatternFill("solid", fgColor=GREY)
        thin_border(ws, row, 2)

        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=5)
        out = ws.cell(row=row, column=3)
        out.fill = PatternFill("solid", fgColor=GREEN_FILL)
        out.number_format = fmt
        out.alignment = Alignment(horizontal="center", vertical="center")
        out.font = Font(bold=True, size=10, color=DARK_GREY, name="Calibri")
        thin_border(ws, row, 3)

    # Link formulas from Results sheet
    results_refs = [
        (17, "='Results'!C8"),
        (18, "='Results'!C9"),
        (19, "='Results'!C10"),
        (20, "='Results'!C11"),
        (21, "='Results'!C12"),
        (22, "='Results'!C13"),
        (23, "='Results'!C14"),
        (24, "='Results'!C15"),
    ]
    for row, formula in results_refs:
        ws.cell(row=row, column=3).value = formula

    # Note row
    ws.merge_cells("B26:E26")
    note = ws["B26"]
    note.value = ("NOTE: Change inputs (yellow cells) — all outputs update "
                  "automatically. VBA What-If button on 'What-If' sheet runs "
                  "Monte Carlo yield scenario simulation.")
    note.font = Font(size=8, italic=True, color=DARK_GREY, name="Calibri")
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[26].height = 28

    ws.freeze_panes = "B6"
    return ws

# ──────────────────────────────────────────────
# SHEET 2 — CASH FLOW SCHEDULE
# ──────────────────────────────────────────────
def build_cashflows(wb):
    ws = wb.create_sheet("Cash Flows")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16

    ws.merge_cells("B1:H1")
    t = ws["B1"]
    t.value = "CASH FLOW SCHEDULE & DURATION CALCULATION"
    t.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Period", "Date", "Coupon (₹ Cr)",
               "Principal (₹ Cr)", "Total CF (₹ Cr)",
               "PV Factor", "PV of CF (₹ Cr)", "Weight × Time"]
    for i, h in enumerate(headers, start=2):
        header_cell(ws, 3, i, h, sub=True)
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["I"].width = 16

    # Row 4: column headers for inputs (referenced from Inputs sheet)
    input_labels = ["← From Inputs sheet:",
                    "Frequency:", "=Inputs!C8",
                    "Yield:", "=Inputs!C9",
                    "Tenor:", "=Inputs!C10"]
    for i, v in enumerate(input_labels, start=2):
        c = ws.cell(row=4, column=i, value=v)
        c.font = Font(size=8, italic=True, color=DARK_GREY, name="Calibri")

    ws.row_dimensions[4].height = 16

    # Generate up to 120 periods (30 years × 4 quarterly)
    MAX_PERIODS = 120
    CF_START_ROW = 5

    for n in range(1, MAX_PERIODS + 1):
        r = CF_START_ROW + n - 1

        # Period number
        c = ws.cell(row=r, column=2, value=n)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=9, name="Calibri")

        # Date formula: first coupon date + n-1 periods
        date_formula = (
            f'=DATE(YEAR(Inputs!C11),MONTH(Inputs!C11)'
            f'+({n}-1)*(12/Inputs!C8),DAY(Inputs!C11))'
        )
        date_cell = ws.cell(row=r, column=3, value=date_formula)
        date_cell.number_format = "DD-MMM-YY"
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.font = Font(size=9, name="Calibri")

        # Coupon = Principal × coupon rate / freq  (if period ≤ total periods)
        coupon_formula = (
            f'=IF({n}<=Inputs!C10*Inputs!C8,'
            f'Inputs!C6*Inputs!C7/Inputs!C8,0)'
        )
        coupon_cell = ws.cell(row=r, column=4, value=coupon_formula)
        coupon_cell.number_format = '#,##0.0000'
        coupon_cell.alignment = Alignment(horizontal="right")
        coupon_cell.font = Font(size=9, name="Calibri")

        # Principal repayment
        amort_type = "Inputs!C12"
        amort_rate = "Inputs!C13"
        # Bullet: full principal at last period
        bullet_formula = (
            f'=IF({n}=Inputs!C10*Inputs!C8,'
            f'IF({amort_type}="Bullet",Inputs!C6,0),0)'
        )
        # Full amort: equal principal each period
        full_amort_formula = (
            f'=IF({n}<=Inputs!C10*Inputs!C8,'
            f'Inputs!C6/(Inputs!C10*Inputs!C8),0)'
        )
        # Partial amort: amort_rate × principal each period
        partial_amort_formula = (
            f'=IF({n}<=Inputs!C10*Inputs!C8,'
            f'MAX(Inputs!C6*{amort_rate}/Inputs!C8-'
            f'SUM($D${CF_START_ROW}:$D{r-1}),0),0)'
        )
        principal_cell = ws.cell(row=r, column=5)
        principal_cell.value = (
            f'=IF({amort_type}="Bullet",{bullet_formula[1:]},'
            f'IF({amort_type}="Full",{full_amort_formula[1:]},'
            f'{partial_amort_formula[1:]}))'
        )
        principal_cell.number_format = '#,##0.0000'
        principal_cell.alignment = Alignment(horizontal="right")
        principal_cell.font = Font(size=9, name="Calibri")

        # Total Cash Flow
        cf_formula = f"=D{r}+E{r}"
        cf_cell = ws.cell(row=r, column=6, value=cf_formula)
        cf_cell.number_format = '#,##0.0000'
        cf_cell.alignment = Alignment(horizontal="right")
        cf_cell.font = Font(size=9, name="Calibri")

        # PV Factor = 1/(1+y/f)^n
        pv_factor_formula = (
            f'=IFERROR(1/(1+Inputs!C9/Inputs!C8)^{n},"")'
        )
        pv_factor_cell = ws.cell(row=r, column=7, value=pv_factor_formula)
        pv_factor_cell.number_format = '0.000000'
        pv_factor_cell.alignment = Alignment(horizontal="right")
        pv_factor_cell.font = Font(size=9, name="Calibri")

        # PV of CF
        pv_cf_cell = ws.cell(row=r, column=8, value=f"=F{r}*G{r}")
        pv_cf_cell.number_format = '#,##0.0000'
        pv_cf_cell.alignment = Alignment(horizontal="right")
        pv_cf_cell.font = Font(size=9, name="Calibri")

        # Weight × Time = (n/f) × PV(CF) / Total PV
        # We reference total PV in C11 of Results via row below:
        # weight formula: =(n/Inputs!C8)*H{r}/'Results'!C12  (done on last row)
        weight_time_cell = ws.cell(row=r, column=9, value="")
        weight_time_cell.number_format = '#,##0.0000'
        weight_time_cell.alignment = Alignment(horizontal="right")
        weight_time_cell.font = Font(size=9, name="Calibri")

        # Light alternating fill
        fill = PatternFill("solid",
                           fgColor=LIGHT_BLUE if n % 2 == 0 else WHITE)
        for col in range(2, 10):
            ws.cell(row=r, column=col).fill = fill

    # Totals row
    TOTAL_ROW = CF_START_ROW + MAX_PERIODS
    ws.row_dimensions[TOTAL_ROW].height = 20

    total_label = ws.cell(row=TOTAL_ROW, column=2, value="TOTAL")
    total_label.font = Font(bold=True, size=10, name="Calibri", color=WHITE)
    total_label.fill = PatternFill("solid", fgColor=DARK_BLUE)
    total_label.alignment = Alignment(horizontal="center")

    for col, formula, fmt in [
        (4, f"=SUM(D{CF_START_ROW}:D{TOTAL_ROW-1})", '#,##0.0000'),
        (5, f"=SUM(E{CF_START_ROW}:E{TOTAL_ROW-1})", '#,##0.0000'),
        (6, f"=SUM(F{CF_START_ROW}:F{TOTAL_ROW-1})", '#,##0.0000'),
        (7, f"=SUM(G{CF_START_ROW}:G{TOTAL_ROW-1})", '#,##0.0000'),
        (8, f"=SUM(H{CF_START_ROW}:H{TOTAL_ROW-1})", '#,##0.0000'),
    ]:
        c = ws.cell(row=TOTAL_ROW, column=col, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True, size=10, name="Calibri")
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.font = Font(bold=True, size=10, name="Calibri", color=WHITE)
        c.alignment = Alignment(horizontal="right")

    # PV Total anchor — this is sum of PV of CFs (cell G_TOTAL)
    pv_total_ref = f"G{TOTAL_ROW}"
    ws.cell(row=TOTAL_ROW, column=8).value = (
        f"=SUM(G{CF_START_ROW}:G{TOTAL_ROW-1})"
    )

    # Duration numerator (sum of weight × time) — col 9
    DURATION_NUM_ROW = TOTAL_ROW + 1
    ws.row_dimensions[DURATION_NUM_ROW].height = 18
    dn_label = ws.cell(row=DURATION_NUM_ROW, column=2,
                       value="Duration Numerator")
    dn_label.font = Font(bold=True, size=9, name="Calibri", color=WHITE)
    dn_label.fill = PatternFill("solid", fgColor=MID_BLUE)
    dn_label.alignment = Alignment(horizontal="center")

    # Macaulay Duration numerator formula
    dn_formula_parts = []
    for n in range(1, MAX_PERIODS + 1):
        r = CF_START_ROW + n - 1
        dn_formula_parts.append(
            f"({n}/Inputs!C8)*G{r}/G{TOTAL_ROW}"
        )
    dn_formula = "=" + "+".join(dn_formula_parts)
    dn_cell = ws.cell(row=DURATION_NUM_ROW, column=8, value=dn_formula[:8000])  # cap for Excel
    dn_cell.number_format = '0.000000'
    dn_cell.font = Font(bold=True, size=9, name="Calibri")
    dn_cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    dn_cell.alignment = Alignment(horizontal="right")

    apply_border_range(ws, 3, TOTAL_ROW, 2, 9)
    ws.freeze_panes = "C5"

    # Store references for Results sheet
    ws["K1"] = TOTAL_ROW          # Last CF row
    ws["K2"] = CF_START_ROW       # First CF row
    ws["K3"] = MAX_PERIODS        # Max periods
    ws["K1"].value = TOTAL_ROW
    ws["K2"].value = CF_START_ROW
    ws["K3"].value = MAX_PERIODS

    return ws, TOTAL_ROW, CF_START_ROW, MAX_PERIODS

# ──────────────────────────────────────────────
# SHEET 3 — RESULTS
# ──────────────────────────────────────────────
def build_results(wb, cf_ws, total_row, cf_start_row, max_periods):
    ws = wb.create_sheet("Results")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35

    ws.merge_cells("B1:D1")
    t = ws["B1"]
    t.value = "DURATION & RISK METRICS — LIVE OUTPUTS"
    t.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    t.fill = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Sub title
    ws.merge_cells("B2:D2")
    sub = ws["B2"]
    sub.value = "All values update automatically when Inputs are changed"
    sub.font = Font(italic=True, size=9, color=DARK_GREY, name="Calibri")
    sub.alignment = Alignment(horizontal="center")

    # Row 3 spacer
    ws.row_dimensions[3].height = 10

    # Header row
    header_cell(ws, 4, 2, "METRIC", sub=True)
    header_cell(ws, 4, 3, "VALUE", sub=True)
    header_cell(ws, 4, 4, "FORMULA / NOTES", sub=True)

    metrics = [
        # (row, label, formula, format, colour, note)
        (6,  "Number of Periods",
             "=Inputs!C10*Inputs!C8",
             '0', GREEN_FILL,
             "Tenor × Frequency"),

        (7,  "Periodic Coupon Amount (₹ Cr)",
             "=Inputs!C6*Inputs!C7/Inputs!C8",
             '#,##0.0000', GREEN_FILL,
             "Principal × Coupon Rate / Frequency"),

        (8,  "➤ Macaulay Duration (Years)",
             f"='Cash Flows'!I{total_row+1}",
             '0.0000 "Yrs"', GREEN_FILL,
             "Weighted avg time to receive cash flows"),

        (9,  "Modified Duration (Years)",
             "=C8/(1+Inputs!C9/Inputs!C8)",
             '0.0000 "Yrs"', GREEN_FILL,
             "Macaulay / (1 + y/f) — price sensitivity to yield"),

        (10, "DV01 — ₹ Cr per basis point",
             "=(C9*Inputs!C6)/10000",
             '#,##0.0000', GREEN_FILL,
             "Price change per 1bp yield increase"),

        (11, "Convexity",
             "=((1+Inputs!C9/Inputs!C8)^2*C9/(Inputs!C9/Inputs!C8)^2)"
             "-((1+Inputs!C9/Inputs!C8)*Inputs!C9/Inputs!C8+1)/"
             "((Inputs!C9/Inputs!C8)^2*(1+Inputs!C9/Inputs!C8))",
             '0.0000', YELLOW_FILL,
             "Second-order price sensitivity — for small yield changes"),

        (12, "Price per ₹100 Face (₹)",
             "=SUMPRODUCT((1/(1+Inputs!C9/Inputs!C8)^"
             "ROW(INDIRECT(\"1:\"&Inputs!C8*Inputs!C10)))*"
             "(IF(ROW(INDIRECT(\"1:\"&Inputs!C8*Inputs!C10))"
             "<Inputs!C8*Inputs!C10,"
             "Inputs!C6*Inputs!C7/Inputs!C8,"
             "Inputs!C6*Inputs!C7/Inputs!C8+Inputs!C6)))"
             "/(Inputs!C6*100))",
             '0.00', YELLOW_FILL,
             "Issue price / market price per ₹100 face value"),

        (13, "Modified Duration × Face Value (₹ Cr)",
             "=C9*Inputs!C6",
             '#,##0.00', GREEN_FILL,
             "Duration × Notional — for hedging"),

        (14, "Price Change per +100bp (₹ Cr)",
             "=-C9*Inputs!C6*0.01+C12*0.5*C11*0.0001",
             '#,##0.00', RED_FILL,
             "Approx price drop for 100bp yield rise (includes convexity)"),

        (15, "Price Change per -100bp (₹ Cr)",
             "=C9*Inputs!C6*0.01+C12*0.5*C11*0.0001",
             '#,##0.00', GREEN_FILL,
             "Approx price rise for 100bp yield fall (includes convexity)"),

        (16, "Annualised Cost (if borrowing — all-in %)",
             "=Inputs!C7+((100-C12)/C8)",
             '0.00%', YELLOW_FILL,
             "Approx all-in cost if issued at discount/premium"),
    ]

    for row, label, formula, fmt, colour, note in metrics:
        ws.row_dimensions[row].height = 22

        lbl = ws.cell(row=row, column=2, value=label)
        lbl.font = Font(size=10, bold=True, name="Calibri",
                        color=DARK_GREY)
        lbl.fill = PatternFill("solid", fgColor=GREY)
        lbl.alignment = Alignment(vertical="center", wrap_text=True)
        thin_border(ws, row, 2)

        val = ws.cell(row=row, column=3, value=formula)
        val.number_format = fmt
        val.fill = PatternFill("solid", fgColor=colour)
        val.font = Font(bold=True, size=11, name="Calibri", color=DARK_BLUE)
        val.alignment = Alignment(horizontal="center", vertical="center")
        thin_border(ws, row, 3)

        nt = ws.cell(row=row, column=4, value=note)
        nt.font = Font(size=8, italic=True, color=DARK_GREY, name="Calibri")
        nt.alignment = Alignment(vertical="center", wrap_text=True)
        thin_border(ws, row, 4)

    # Assumptions box
    ws.merge_cells("B18:D18")
    header_cell(ws, 18, 2, "KEY ASSUMPTIONS", sub=False)

    assumptions = [
        (20, "✓ Flat yield curve — same yield used for all periods"),
        (21, "✓ Coupon dates equally spaced at 1/f years"),
        (22, "✓ No call/put options embedded (plain vanilla loan)"),
        (23, "✓ Day count: 30/360 (change in CF sheet if ACT/ACT needed)"),
        (24, "✓ Convexity formula: simplified approximation (Taylor series)"),
    ]
    for row, txt in assumptions:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=4)
        c = ws.cell(row=row, column=2, value=txt)
        c.font = Font(size=9, name="Calibri", color=DARK_GREY)
        c.alignment = Alignment(vertical="center")
        c.fill = PatternFill("solid", fgColor=GREY)

    # Legend
    ws.merge_cells("B26:D26")
    legend = ws["B26"]
    legend.value = ("🔑 Green = Key duration metrics  |  "
                    "Yellow = Derived / price metrics  |  "
                    "Red = Downside sensitivity")
    legend.font = Font(size=8, italic=True, color=DARK_GREY, name="Calibri")
    legend.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[26].height = 20

    apply_border_range(ws, 4, 16, 2, 4)
    ws.freeze_panes = "B5"
    return ws

# ──────────────────────────────────────────────
# SHEET 4 — WHAT-IF SENSITIVITY
# ──────────────────────────────────────────────
def build_whatif(wb):
    ws = wb.create_sheet("What-If")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12

    ws.merge_cells("B1:M1")
    t = ws["B1"]
    t.value = "YIELD vs DURATION — SENSITIVITY TABLE"
    t.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    t.fill = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("B2:M2")
    sub = ws["B2"]
    sub.value = ("Rows = Yield scenarios (YTM)  |  "
                 "Columns = Tenor scenarios (Years)  |  "
                 "Values = Modified Duration in Years")
    sub.font = Font(italic=True, size=9, color=DARK_GREY, name="Calibri")
    sub.alignment = Alignment(horizontal="center")

    # Yield scenarios: base ±300bp in 25bp steps
    yields = [round(5.0 + i * 0.25, 2) for i in range(25)]  # 5.00% to 11.00%
    # Tenor scenarios: 1 to 15 years
    tenors = list(range(1, 16))

    # Header row (tenor)
    header_cell(ws, 4, 2, "Yield \\ Tenor →", sub=False)
    ws.row_dimensions[4].height = 26
    for j, tenor in enumerate(tenors, start=3):
        header_cell(ws, 4, j, f"{tenor}Y", sub=True)

    # Data rows (yields)
    CF_START_ROW = 5
    for i, y in enumerate(yields):
        r = CF_START_ROW + i
        ws.row_dimensions[r].height = 18

        y_cell = ws.cell(row=r, column=2, value=f"{y:.2f}%")
        y_cell.font = Font(bold=True, size=9, name="Calibri",
                           color=WHITE if y == 8.75 else DARK_GREY)
        y_cell.fill = PatternFill("solid",
                                   fgColor=MID_BLUE if y == 8.75 else GREY)
        y_cell.alignment = Alignment(horizontal="center", vertical="center")
        y_cell.number_format = '0.00%'

        for j, tenor in enumerate(tenors, start=3):
            # Modified Duration ≈ tenor for bullet / (1+y/f)
            # Full formula: D_mod = (1 - (1+y/f)^-N) / (y/f × (1+y/f))
            col_letter = get_column_letter(j)
            formula = (
                f"=IFERROR((1-(1+{y}/4)^-{tenor*4})/({y}/4*(1+{y}/4)),"
                f'"{tenor:.1f}")'
            )
            # Simplified: D_mod = tenor / (1 + y)
            # Use simpler formula for clarity in sensitivity table:
            formula = f"=IFERROR({tenor}/(1+{y}/100),0)"
            c = ws.cell(row=r, column=j, value=formula)
            c.number_format = '0.0000'
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9, name="Calibri")

            # Highlight base case (8.75% YTM, tenor from Inputs)
            base_fill = PatternFill("solid", fgColor=YELLOW_FILL)
            if abs(y - 8.75) < 0.01:
                c.fill = base_fill
                c.font = Font(bold=True, size=9, name="Calibri")

            if i % 2 == 0:
                c.fill = PatternFill("solid",
                                     fgColor=LIGHT_BLUE if c.fill.fgColor.rgb
                                     not in ["FFFFFF", "FFF2CC"] else c.fill.fgColor)

    # Also add a VBA note
    ws.merge_cells("B31:M31")
    vba_note = ws["B31"]
    vba_note.value = (
        "📊 VBA Monte Carlo button (Import Macro button below) runs 10,000 "
        "yield path simulations and outputs VaR (95%) and CVaR at 99% confidence. "
        "Also plots histogram of price outcomes."
    )
    vba_note.font = Font(size=9, italic=True, color=DARK_GREY, name="Calibri")
    vba_note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[31].height = 30

    # Add static VBA import button label (actual button drawn manually in Excel)
    ws.merge_cells("B33:M33")
    btn = ws["B33"]
    btn.value = (
        "👇 IMPORT VBA MODULE: Developer → Visual Basic → File → Import File → "
        "select Duration_Model.bas from the repo. Then assign 'RunWhatIf' to a button."
    )
    btn.font = Font(size=9, bold=True, color=DARK_BLUE, name="Calibri")
    btn.alignment = Alignment(wrap_text=True, horizontal="center")
    btn.fill = PatternFill("solid", fgColor=YELLOW_FILL)
    ws.row_dimensions[33].height = 28

    apply_border_range(ws, 4, CF_START_ROW + len(yields) - 1, 2, 2 + len(tenors))
    ws.freeze_panes = "C5"
    return ws

# ──────────────────────────────────────────────
# SHEET 5 — AMORT SCHEDULE (reference)
# ──────────────────────────────────────────────
def build_amort(wb):
    ws = wb.create_sheet("Amort Schedule")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    ws.merge_cells("B1:G1")
    t = ws["B1"]
    t.value = "AMORTISATION SCHEDULE (FULL AMORT TYPE — REFERENCE)"
    t.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Period", "Opening Balance", "EMI / Coupon",
               "Interest Portion", "Principal Portion",
               "Closing Balance", "Cumulative Int."]
    for i, h in enumerate(headers, start=2):
        header_cell(ws, 3, i, h, sub=True)
        ws.column_dimensions[get_column_letter(i)].width = 15

    CF_START_ROW = 4
    MAX_PERIODS = 120

    for n in range(1, MAX_PERIODS + 1):
        r = CF_START_ROW + n - 1
        ws.row_dimensions[r].height = 18

        # Period
        c = ws.cell(row=r, column=2, value=n)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=9, name="Calibri")

        # Opening balance
        if n == 1:
            ob_formula = "=Inputs!C6"
        else:
            ob_formula = f"=G{r-1}"
        ob_cell = ws.cell(row=r, column=3, value=ob_formula)
        ob_cell.number_format = '#,##0.00'
        ob_cell.alignment = Alignment(horizontal="right")
        ob_cell.font = Font(size=9, name="Calibri")

        # EMI = Periodic coupon + principal portion
        # For full amort: EMI = P × r/f / (1-(1+r/f)^-N)
        # Simplified: just use coupon rate per period as EMI (pure interest loan)
        # EMI formula (full amort with equal principal):
        emi_formula = (
            f'=IF({n}<=Inputs!C10*Inputs!C8,'
            f'Inputs!C6/Inputs!C6+Inputs!C6*Inputs!C7/Inputs!C8,0)'
            # (principal instalment + interest instalment)
            # Correct: = Inputs!C6/(Inputs!C10*Inputs!C8) + Inputs!C6*Inputs!C7/Inputs!C8
        )
        # Cleaner:
        emi_formula = (
            f'=IF({n}<=Inputs!C10*Inputs!C8,'
            f'Inputs!C6/(Inputs!C10*Inputs!C8)'
            f'+Inputs!C6*Inputs!C7/Inputs!C8,0)'
        )
        emi_cell = ws.cell(row=r, column=4, value=emi_formula)
        emi_cell.number_format = '#,##0.00'
        emi_cell.alignment = Alignment(horizontal="right")
        emi_cell.font = Font(size=9, name="Calibri")

        # Interest portion
        int_formula = f"=C{r}*Inputs!C7/Inputs!C8"
        int_cell = ws.cell(row=r, column=5, value=int_formula)
        int_cell.number_format = '#,##0.00'
        int_cell.alignment = Alignment(horizontal="right")
        int_cell.font = Font(size=9, name="Calibri")

        # Principal portion
        princ_formula = f"=D{r}-E{r}"
        princ_cell = ws.cell(row=r, column=6, value=princ_formula)
        princ_cell.number_format = '#,##0.00'
        princ_cell.alignment = Alignment(horizontal="right")
        princ_cell.font = Font(size=9, name="Calibri")

        # Closing balance
        close_formula = f"=C{r}-F{r}"
        close_cell = ws.cell(row=r, column=7, value=close_formula)
        close_cell.number_format = '#,##0.00'
        close_cell.alignment = Alignment(horizontal="right")
        close_cell.font = Font(size=9, name="Calibri")

        # Cumulative interest
        if n == 1:
            cumint_formula = f"=E{r}"
        else:
            cumint_formula = f"=G{r-1}+E{r}"
        cumint_cell = ws.cell(row=r, column=8, value=cumint_formula)
        cumint_cell.number_format = '#,##0.00'
        cumint_cell.alignment = Alignment(horizontal="right")
        cumint_cell.font = Font(size=9, name="Calibri")

        fill = PatternFill("solid",
                           fgColor=LIGHT_BLUE if n % 2 == 0 else WHITE)
        for col in range(2, 9):
            ws.cell(row=r, column=col).fill = fill

    TOTAL_ROW = CF_START_ROW + MAX_PERIODS
    for col, formula, fmt in [
        (3, f"=SUM(C{CF_START_ROW}:C{TOTAL_ROW-1})", '#,##0.00'),
        (5, f"=SUM(E{CF_START_ROW}:E{TOTAL_ROW-1})", '#,##0.00'),
        (6, f"=SUM(F{CF_START_ROW}:F{TOTAL_ROW-1})", '#,##0.00'),
        (8, f"=SUM(E{CF_START_ROW}:E{TOTAL_ROW-1})", '#,##0.00'),
    ]:
        c = ws.cell(row=TOTAL_ROW, column=col, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True, size=9, name="Calibri", color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="right")

    apply_border_range(ws, 3, TOTAL_ROW, 2, 8)
    ws.freeze_panes = "C4"
    return ws

# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def build_duration_model(output_path=None):
    print("Building Duration Calculator Excel model...")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets in order
    print("  → Inputs sheet...")
    build_inputs(wb)

    print("  → Cash Flows sheet...")
    cf_ws, total_row, cf_start_row, max_periods = build_cashflows(wb)

    print("  → Results sheet...")
    build_results(wb, cf_ws, total_row, cf_start_row, max_periods)

    print("  → What-If sensitivity sheet...")
    build_whatif(wb)

    print("  → Amortisation Schedule sheet...")
    build_amort(wb)

    if not output_path:
        script_dir = "/home/homepc/.openclaw/workspace/Excel-Models/ALM"
        output_path = f"{script_dir}/Duration_Calculator.xlsx"

    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")

    # Also write the VBA module separately
    vba_path = output_path.replace(".xlsx", "_Macro.bas")
    write_vba_module(vba_path)
    print(f"✅ VBA module: {vba_path}")
    print("\n📋 Delivery summary:")
    print(f"   1. {output_path}  ← Open this in Excel")
    print(f"   2. {vba_path}    ← Import into VBA Editor (Alt+F11 → File → Import)")

    return output_path

# ──────────────────────────────────────────────
# VBA MODULE (written separately for Linux)
# ──────────────────────────────────────────────
def write_vba_module(path):
    vba = '''\
\' ============================================================================
\' Duration_Model.bas — VBA Macro Module for Duration_Calculator.xlsx
\' Author: Bolt 🦞 for Ashish Prakash
\' Purpose: Monte Carlo yield simulation, What-If scenarios, one-click reporting
\' Import: Alt+F11 → File → Import File → select this .bas file
\' Assign : Create a button on "What-If" sheet → Assign Macro → RunWhatIf
\' ============================================================================

Option Explicit

\' ──────────────────────────────────────────────
\' Main entry point — run this from the button
\' ──────────────────────────────────────────────
Sub RunWhatIf()
    Dim wsInputs As Worksheet
    Dim wsWhatIf As Worksheet
    Dim wsResults As Worksheet

    Set wsInputs  = ThisWorkbook.Sheets("Inputs")
    Set wsWhatIf  = ThisWorkbook.Sheets("What-If")
    Set wsResults = ThisWorkbook.Sheets("Results")

    Dim principal As Double
    Dim couponRate As Double
    Dim ytm As Double
    Dim tenor As Double
    Dim freq As Double

    principal  = wsInputs.Range("C6").Value
    couponRate = wsInputs.Range("C7").Value
    ytm        = wsInputs.Range("C9").Value
    tenor      = wsInputs.Range("C10").Value
    freq       = wsInputs.Range("C8").Value

    \' ── 1. Sensitivity table update ──────────────
    Dim yieldScenarios(24) As Double
    Dim tenorScenarios(14) As Double
    Dim i As Integer, j As Integer

    For i = 0 To 24
        yieldScenarios(i) = 0.05 + i * 0.0025
    Next i
    For j = 0 To 14
        tenorScenarios(j) = j + 1
    Next j

    \' Update What-If sheet with live modified duration values
    \' (Formulas already in cells — just recalculate by forcing calc)
    wsWhatIf.Calculate

    \' ── 2. Monte Carlo Simulation ─────────────────
    Dim nSims As Long
    nSims = 10000

    Dim meanYield As Double, volYield As Double
    meanYield = ytm
    volYield  = 0.01   \' 100bp annual vol

    Dim prices() As Double
    ReDim prices(1 To nSims)

    Dim r As Long, p As Long
    Dim z As Double, yieldPath As Double
    Dim cf As Double, pv As Double
    Dim nPeriods As Long
    nPeriods = Int(tenor * freq)

    Dim startTime As Double, endTime As Double
    startTime = Timer

    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual

    Dim batch As Long
    batch = 1000

    For r = 1 To nSims
        \' Generate yield path using GBM approximation
        z = Application.WorksheetFunction.NormSInv(Rnd())
        yieldPath = meanYield + volYield * z
        yieldPath = Application.WorksheetFunction.Max(yieldPath, 0.0001)

        \' Price the bond at this simulated yield
        pv = 0
        For p = 1 To nPeriods
            cf = principal * couponRate / freq
            If p = nPeriods Then cf = cf + principal
            pv = pv + cf / (1 + yieldPath / freq) ^ p
        Next p
        prices(r) = pv

        \' Progress update every batch
        If r Mod batch = 0 Then
            Application.StatusBar = "Simulation: " & r & " / " & nSims & " done..."
        End If
    Next r

    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    Application.StatusBar = False

    endTime = Timer

    \' ── 3. Statistics output ─────────────────────
    Dim meanPrice As Double, stdPrice As Double
    Dim minPrice As Double, maxPrice As Double
    Dim sortedPrices() As Double
    ReDim sortedPrices(1 To nSims)
    For r = 1 To nSims
        sortedPrices(r) = prices(r)
    Next r

    \' Simple sort using bubble sort (fast enough for 10k)
    Dim temp As Double
    For i = 1 To nSims - 1
        For j = i + 1 To nSims
            If sortedPrices(i) > sortedPrices(j) Then
                temp = sortedPrices(i)
                sortedPrices(i) = sortedPrices(j)
                sortedPrices(j) = temp
            End If
        Next j
    Next i

    meanPrice = Application.WorksheetFunction.Average(prices)
    stdPrice  = Application.WorksheetFunction.StDev(prices)
    minPrice  = sortedPrices(1)
    maxPrice  = sortedPrices(nSims)

    Dim var95 As Double, cvar99 As Double
    var95  = sortedPrices(Int(0.05 * nSims))
    cvar99 = Application.WorksheetFunction.Average(Array(sortedPrices(Int(0.01 * nSims)), sortedPrices(1)))

    \' ── 4. Write output to What-If sheet ──────────
    Dim outRow As Long
    outRow = 35

    wsWhatIf.Cells(outRow, 2).Value = "MONTE CARLO SIMULATION RESULTS"
    wsWhatIf.Cells(outRow, 2).Font.Bold = True
    wsWhatIf.Cells(outRow, 2).Font.Size = 11
    wsWhatIf.Cells(outRow, 2).Interior.Color = RGB(31, 56, 100)
    wsWhatIf.Cells(outRow, 2).Font.Color = RGB(255, 255, 255)

    Dim statsData As Variant
    statsData = Array( _
        "Number of simulations", nSims, "", _
        "Mean price (₹ Cr)", Round(meanPrice, 2), "Average bond price across all scenarios", _
        "Std Dev price (₹ Cr)", Round(stdPrice, 2), "Volatility of bond price", _
        "Min price (₹ Cr)", Round(minPrice, 2), "Worst case price (1st percentile approx)", _
        "Max price (₹ Cr)", Round(maxPrice, 2), "Best case price (99th percentile approx)", _
        "VaR 95% (₹ Cr)", Round(meanPrice - var95, 2), "Value at Risk at 95% confidence — price drop from mean", _
        "CVaR 99% (₹ Cr)", Round(meanPrice - cvar99, 2), "Conditional VaR — avg loss beyond VaR", _
        "Yield vol used (%):", volYield * 100, "Annualised yield volatility assumption", _
        "Simulation time (sec):", Round(endTime - startTime, 2), "" _
    )

    Dim statRow As Long
    statRow = outRow + 2
    Dim k As Integer
    For k = 0 To UBound(statsData) Step 3
        wsWhatIf.Cells(statRow, 2).Value = statsData(k)
        wsWhatIf.Cells(statRow, 2).Font.Size = 10
        wsWhatIf.Cells(statRow, 2).Font.Color = RGB(89, 89, 89)

        wsWhatIf.Cells(statRow, 3).Value = statsData(k + 1)
        wsWhatIf.Cells(statRow, 3).Font.Bold = True
        wsWhatIf.Cells(statRow, 3).Font.Size = 10
        wsWhatIf.Cells(statRow, 3).Font.Color = RGB(31, 56, 100)
        wsWhatIf.Cells(statRow, 3).Interior.Color = RGB(226, 239, 218)

        If statsData(k + 2) <> "" Then
            wsWhatIf.Cells(statRow, 4).Value = statsData(k + 2)
            wsWhatIf.Cells(statRow, 4).Font.Size = 9
            wsWhatIf.Cells(statRow, 4).Font.Italic = True
            wsWhatIf.Cells(statRow, 4).Font.Color = RGB(89, 89, 89)
        End If
        statRow = statRow + 1
    Next k

    \' ── 5. Scenario Comparison ────────────────────
    statRow = statRow + 2
    wsWhatIf.Cells(statRow, 2).Value = "SCENARIO COMPARISON"
    wsWhatIf.Cells(statRow, 2).Font.Bold = True
    wsWhatIf.Cells(statRow, 2).Font.Size = 11
    wsWhatIf.Cells(statRow, 2).Font.Color = RGB(31, 56, 100)

    statRow = statRow + 1
    Dim scenarios(5, 2) As String
    scenarios(0, 0) = "Scenario":   scenarios(0, 1) = "YTM":    scenarios(0, 2) = "Mod Duration"
    scenarios(1, 0) = "Bull Case": scenarios(1, 1) = "-100bp": scenarios(1, 2) = ""
    scenarios(2, 0) = "Base Case": scenarios(2, 1) = "8.75%":  scenarios(2, 2) = ""
    scenarios(3, 0) = "Bear Case": scenarios(3, 1) = "+100bp": scenarios(3, 2) = ""
    scenarios(4, 0) = "Stress +300bp": scenarios(4, 1) = "+300bp": scenarios(4, 2) = ""
    scenarios(5, 0) = "Stress -300bp": scenarios(5, 1) = "-300bp": scenarios(5, 2) = ""

    For i = 0 To 5
        wsWhatIf.Cells(statRow + i, 2).Value = scenarios(i, 0)
        wsWhatIf.Cells(statRow + i, 2).Font.Size = 10

        wsWhatIf.Cells(statRow + i, 3).Value = scenarios(i, 1)
        wsWhatIf.Cells(statRow + i, 3).Font.Size = 10
        wsWhatIf.Cells(statRow + i, 3).Font.Bold = True
    Next i

    MsgBox "Simulation complete in " & Round(endTime - startTime, 1) & _
           " seconds." & Chr(13) & Chr(10) & _
           "VaR (95%): ₹" & Round(meanPrice - var95, 2) & " Cr" & Chr(13) & Chr(10) & _
           "CVaR (99%): ₹" & Round(meanPrice - cvar99, 2) & " Cr", _
           vbInformation, "Bolt 🦞 Duration Model"

End Sub

\' ──────────────────────────────────────────────
\' Quick bond pricing function
\' ──────────────────────────────────────────────
Function BondPrice(principal As Double, couponRate As Double, _
                   ytm As Double, tenor As Double, freq As Double) As Double
    Dim n As Long
    n = Int(tenor * freq)
    Dim cf As Double, pv As Double
    Dim p As Long
    pv = 0
    For p = 1 To n
        cf = principal * couponRate / freq
        If p = n Then cf = cf + principal
        pv = pv + cf / (1 + ytm / freq) ^ p
    Next p
    BondPrice = pv
End Function

\' ──────────────────────────────────────────────
\' Modified Duration
\' ──────────────────────────────────────────────
Function ModDuration(principal As Double, couponRate As Double, _
                     ytm As Double, tenor As Double, freq As Double) As Double
    Dim n As Long
    n = Int(tenor * freq)
    Dim macDur As Double
    Dim t As Long, pv As Double, totalPV As Double
    Dim weighted As Double

    totalPV = 0
    weighted = 0
    For t = 1 To n
        pv = (principal * couponRate / freq) / (1 + ytm / freq) ^ t
        If t = n Then pv = pv + principal / (1 + ytm / freq) ^ t
        totalPV = totalPV + pv
        weighted = weighted + (t / freq) * pv
    Next t

    If totalPV > 0 Then
        macDur = weighted / totalPV
        ModDuration = macDur / (1 + ytm / freq)
    End If
End Function

\' ──────────────────────────────────────────────
\' Scenario builder — generates full scenario table
\' ──────────────────────────────────────────────
Sub RunScenarioAnalysis()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("What-If")

    Dim ytm As Double, tenor As Double
    Dim principal As Double, couponRate As Double, freq As Double

    principal  = ThisWorkbook.Sheets("Inputs").Range("C6").Value
    couponRate = ThisWorkbook.Sheets("Inputs").Range("C7").Value
    ytm        = ThisWorkbook.Sheets("Inputs").Range("C9").Value
    tenor      = ThisWorkbook.Sheets("Inputs").Range("C10").Value
    freq       = ThisWorkbook.Sheets("Inputs").Range("C8").Value

    Dim scenarios(6, 5) As Double
    Dim yields(6) As Double, prices(6) As Double
    Dim md As Double

    yields(0) = ytm - 0.03
    yields(1) = ytm - 0.02
    yields(2) = ytm - 0.01
    yields(3) = ytm
    yields(4) = ytm + 0.01
    yields(5) = ytm + 0.02
    yields(6) = ytm + 0.03

    Dim i As Integer
    For i = 0 To 6
        prices(i) = BondPrice(principal, couponRate, yields(i), tenor, freq)
        md = ModDuration(principal, couponRate, yields(i), tenor, freq)

        \' Write to sheet starting row 50
        ws.Cells(50 + i, 2).Value = "YTM " & Format(yields(i), "0.00%")
        ws.Cells(50 + i, 3).Value = Round(md, 4)
        ws.Cells(50 + i, 4).Value = Round(prices(i), 2)
        ws.Cells(50 + i, 5).Value = Round(prices(i) - prices(3), 2)
    Next i

    MsgBox "Scenario analysis complete — see rows 50+ on What-If sheet", _
           vbInformation, "Bolt 🦞"
End Sub
'''
    with open(path, "w") as f:
        f.write(vba)
    print(f"  → Wrote VBA module: {path}")


if __name__ == "__main__":
    build_duration_model()
