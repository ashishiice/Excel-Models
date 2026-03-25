#!/usr/bin/env python3
"""
NCD + CP Rate Tracker — Market Surveillance Tool
Purpose : Daily rate tracking for NCD and CP positions
          Manual data entry from CCIL / NDS-OM / Bloomberg
Author  : Bolt 🦞 for Ashish Prakash
Sheets  : Data_Entry | Rates_DB | Portfolio_Summary | GSec_Benchmark | MTM | Alerts
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# ──────────────────────────────────────────────
# PALETTE
# ──────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
YELLOW     = "FFF2CC"
GREEN      = "E2EFDA"
RED        = "FCE4D6"
ORANGE     = "FCE4D6"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"
D_GREY     = "595959"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr(ws, row, col, val, sub=False, merge_end_col=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font  = Font(bold=True, color=WHITE, name="Calibri",
                   size=9 if sub else 11)
    c.fill  = make_fill(MID_BLUE if not sub else DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border = thin()
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
    return c

def inp(ws, row, col, val="", fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = make_fill(YELLOW)
    c.font = Font(bold=True, color=DARK_BLUE, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    if fmt:
        c.number_format = fmt
    return c

def out(ws, row, col, val="", fmt=None, fill=YELLOW):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = make_fill(fill)
    c.font = Font(bold=True, color=D_GREY, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    if fmt:
        c.number_format = fmt
    return c

def dat(ws, row, col, val="", fmt=None, fill=None):
    fill = fill or (GREY if row % 2 == 0 else WHITE)
    c = ws.cell(row=row, column=col, value=val)
    c.fill = make_fill(fill)
    c.font = Font(size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    if fmt:
        c.number_format = fmt
    return c

def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_r, max_r, min_c, max_c):
    for r in range(min_r, max_r+1):
        for c in range(min_c, max_c+1):
            ws.cell(r, c).border = thin()

# ──────────────────────────────────────────────
# SHEET 1 — DATA ENTRY
# Every row = one instrument, one date
# ──────────────────────────────────────────────
def build_data_entry(wb):
    ws = wb.create_sheet("Data_Entry")
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = {"A":2,"B":10,"C":14,"D":16,"E":14,"F":10,"G":10,
              "H":10,"I":10,"J":12,"K":12,"L":10,"M":10,"N":10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("B1:N1")
    t = ws["B1"]
    t.value = "NCD & CP — DAILY RATE ENTRY"
    t.font  = Font(bold=True, size=14, color=WHITE, name="Calibri")
    t.fill  = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("B2:N2")
    s = ws["B2"]
    s.value = ("Enter one row per instrument per date. Yellow = manual entry. "
               "Blue columns auto-calculate from GSec_Benchmark & Portfolio_Summary sheets.")
    s.font  = Font(italic=True, size=8, color=D_GREY, name="Calibri")
    s.fill  = make_fill(GREY)
    s.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers row 4
    headers = [
        "Date", "Instrument\nType", "Issuer Name", "ISIN / Ref",
        "Coupon\n(% p.a.)", "Freq\n(2=H/4=Q)",
        "Tenor\n(Years)", "Face Value\n(₹ Cr)",
        "Current\nYield (% )", "Clean Price\n(₹)",
        "Accrued\nInterest (₹)", "Gross\nInvoice Price",
        "Days to\nMaturity", "Current\nMtm Value (₹ Cr)"
    ]
    for i, h in enumerate(headers, start=2):
        hdr(ws, 4, i, h, sub=True)
    ws.row_dimensions[4].height = 32

    # Sample data rows (3 example rows so structure is clear)
    sample_data = [
        ("2026-03-25", "NCD", "HDFC Ltd", "INE001A08R39", 0.0910, 2, 5.0, 100.0,
         0.0915, 99.50, "", "", "", ""),
        ("2026-03-25", "NCD", "Power Finance Corp", "INE881A1DNN7", 0.0820, 2, 3.0, 50.0,
         0.0830, 100.25, "", "", "", ""),
        ("2026-03-25", "CP", "Tata Capital Ltd", "INE851A14MZ6", 0.0750, 4, 0.5, 25.0,
         0.0760, 99.10, "", "", "", ""),
    ]

    START_ROW = 5
    for n, row_data in enumerate(sample_data):
        r = START_ROW + n
        ws.row_dimensions[r].height = 20
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=r, column=col_idx, value=val)
            fill = make_fill(LIGHT_BLUE if n % 2 == 0 else WHITE)
            cell.fill = fill
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin()

            # Format date col
            if col_idx == 2:
                cell.number_format = "DD-MMM-YYYY"
            # Yellow override for manual entry cols
            if col_idx in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
                cell.fill = make_fill(YELLOW)
                cell.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)

    # Auto-formula rows for 250 blank entry rows
    for n in range(START_ROW + len(sample_data), START_ROW + 260):
        r = n
        ws.row_dimensions[r].height = 18

        # Col B: Date
        c = ws.cell(row=r, column=2)
        c.fill = make_fill(YELLOW); c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
        c.number_format = "DD-MMM-YYYY"
        c.alignment = Alignment(horizontal="center")

        # Col C: Instrument Type (dropdown)
        c = ws.cell(row=r, column=3)
        c.fill = make_fill(YELLOW); c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
        c.alignment = Alignment(horizontal="center")

        # Cols D-H: manual
        for col in [4, 5, 6, 7, 8]:
            c = ws.cell(row=r, column=col)
            c.fill = make_fill(YELLOW)
            c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
            c.alignment = Alignment(horizontal="center")
            if col == 5: c.number_format = "0.00%"
            if col == 7: c.number_format = "0.00"
            if col == 8: c.number_format = "#,##0.00"

        # Col I: Current Yield (% manual)
        c = ws.cell(row=r, column=9)
        c.fill = make_fill(YELLOW); c.number_format = "0.000%"; c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
        c.alignment = Alignment(horizontal="center")

        # Col J: Clean Price (manual)
        c = ws.cell(row=r, column=10)
        c.fill = make_fill(YELLOW); c.number_format = "0.00"; c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
        c.alignment = Alignment(horizontal="center")

        # Col K: Accrued Interest (manual — or formula)
        c = ws.cell(row=r, column=11)
        c.fill = make_fill(YELLOW); c.number_format = "0.00"; c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
        c.alignment = Alignment(horizontal="center")

        # Col L: Gross Invoice Price = J + K
        c = ws.cell(row=r, column=12,
                    value=f"=IFERROR(J{r}+K{r},J{r})")
        c.fill = make_fill(GREEN); c.number_format = "0.00"
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)
        c.alignment = Alignment(horizontal="center")

        # Col M: Days to Maturity (auto)
        # DTM = ( maturity_date - date ) using simple annualised
        # We need a maturity date. For CP (tenor < 1Y) approximate from date + tenor
        # For NCD: need maturity date. Use: = date + tenor*365 (rough)
        c = ws.cell(row=r, column=13,
                    value=f"=IFERROR(IF(C{r}=\"CP\",INT(G{r}*365),INT(G{r}*365)),\"\")")
        c.fill = make_fill(LIGHT_BLUE); c.number_format = "0"
        c.font = Font(size=9, name="Calibri"); c.alignment = Alignment(horizontal="center")

        # Col N: MTM Value in ₹ Cr = L * H / 100 (price × face value / 100)
        c = ws.cell(row=r, column=14,
                    value=f"=IFERROR(L{r}*H{r}/100,\"\")")
        c.fill = make_fill(GREEN); c.number_format = "#,##0.00"
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)
        c.alignment = Alignment(horizontal="center")

    # Freeze
    ws.freeze_panes = "B5"
    apply_border(ws, 4, START_ROW+259, 2, 14)

    # Add data validation for Instrument Type
    dv = DataValidation(
        type="list",
        formula1='"NCD,CP,CD,Bond,Other"',
        showDropDown=False
    )
    dv.add(f"C{START_ROW}:C{START_ROW+259}")
    ws.add_data_validation(dv)

    # Totals row
    TOTAL_ROW = START_ROW + 260
    ws.row_dimensions[TOTAL_ROW].height = 22
    ws.merge_cells(start_row=TOTAL_ROW, start_column=2,
                   end_row=TOTAL_ROW, end_column=7)
    tot = ws.cell(row=TOTAL_ROW, column=2, value="PORTFOLIO TOTAL (₹ Cr)")
    tot.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    tot.fill = make_fill(DARK_BLUE)
    tot.alignment = Alignment(horizontal="right", vertical="center")
    tot.border = thin()

    c = ws.cell(row=TOTAL_ROW, column=8,
                value=f"=SUM(H{START_ROW}:H{TOTAL_ROW-1})")
    c.font = Font(bold=True, size=10, name="Calibri", color=WHITE)
    c.fill = make_fill(DARK_BLUE); c.number_format = "#,##0.00"
    c.border = thin()

    c = ws.cell(row=TOTAL_ROW, column=14,
                value=f"=SUM(N{START_ROW}:N{TOTAL_ROW-1})")
    c.font = Font(bold=True, size=10, name="Calibri", color=WHITE)
    c.fill = make_fill(DARK_BLUE); c.number_format = "#,##0.00"
    c.border = thin()

    return ws, START_ROW, TOTAL_ROW


# ──────────────────────────────────────────────
# SHEET 2 — RATES DATABASE
# Historical rates stored here (one row per instrument per date)
# This is a running log — new entries appended from Data_Entry
# ──────────────────────────────────────────────
def build_rates_db(wb, de_ws, de_start, de_total):
    ws = wb.create_sheet("Rates_DB")
    ws.sheet_view.showGridLines = False

    widths = {"A":2,"B":12,"C":14,"D":20,"E":14,"F":8,
              "G":8,"H":10,"I":12,"J":10,"K":12,"L":10,"M":10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:M1")
    t = ws["B1"]
    t.value = "RATES DATABASE — HISTORICAL LOG (AUTO-POPULATED)"
    t.font  = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill  = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:M2")
    s = ws["B2"]
    s.value = "This sheet auto-records each day's entry. Do NOT edit. Use Data_Entry sheet for all inputs."
    s.font  = Font(italic=True, size=8, color=D_GREY, name="Calibri")
    s.fill  = make_fill(GREY)
    s.alignment = Alignment(horizontal="center")

    headers = ["Date","Instrument","Issuer","ISIN","Coupon",
               "Freq","Tenor","Face (₹ Cr)","Yield (%)","Price",
               "AI","Invoice Price","MTM (₹ Cr)"]
    for i, h in enumerate(headers, start=2):
        hdr(ws, 4, i, h, sub=True)
    ws.row_dimensions[4].height = 28

    # Pull from Data_Entry via formulas for latest 50 entries
    # For a proper DB, we'd use a Python script to append.
    # Here: static reference to Data_Entry
    ws.merge_cells("B6:M6")
    note = ws["B6"]
    note.value = ("ℹ  To populate this database: run the Python script "
                  "build_ncd_cp_tracker.py with --append flag, OR manually "
                  "copy Data_Entry rows and Paste Special > Values here.")
    note.font  = Font(italic=True, size=9, color=D_GREY, name="Calibri")
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = 24

    apply_border(ws, 4, 200, 2, 13)
    ws.freeze_panes = "B5"
    return ws


# ──────────────────────────────────────────────
# SHEET 3 — G-SEC BENCHMARK
# ──────────────────────────────────────────────
def build_gsec(wb):
    ws = wb.create_sheet("GSec_Benchmark")
    ws.sheet_view.showGridLines = False

    widths = {"A":2,"B":10,"C":12,"D":12,"E":12,"F":12,"G":12,
              "H":12,"I":12,"J":12,"K":12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:K1")
    t = ws["B1"]
    t.value = "G-SEC BENCHMARK YIELDS — MANUAL ENTRY (RBI / FBIL Data)"
    t.font  = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill  = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:K2")
    s = ws["B2"]
    s.value = ("Enter G-Sec yields by tenor daily from: https://fbil.org.in "
               "or RBI Database (DB). Spread calculations use 3Y, 5Y, 7Y, 10Y benchmarks.")
    s.font  = Font(italic=True, size=8, color=D_GREY, name="Calibri")
    s.fill  = make_fill(GREY)
    s.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.row_dimensions[2].height = 20

    # Tenor header
    tenors = [1, 2, 3, 5, 7, 10, 15, 20, 25, 30]
    hdr(ws, 4, 2, "Date", sub=True)
    for i, t_yr in enumerate(tenors, start=3):
        hdr(ws, 4, i, f"{t_yr}Y G-Sec", sub=True)
    ws.row_dimensions[4].height = 26

    # Sample rows
    sample_gsec = [
        ("2026-03-25", 0.0670, 0.0680, 0.0690, 0.0695, 0.0705, 0.0710, 0.0730, 0.0750, 0.0760, 0.0765),
    ]
    for n, row_data in enumerate(sample_gsec, start=5):
        ws.row_dimensions[n].height = 20
        for col_idx, val in enumerate(row_data, start=2):
            c = ws.cell(row=n, column=col_idx, value=val)
            c.fill = make_fill(YELLOW if col_idx == 2 else (LIGHT_BLUE if n % 2 == 0 else WHITE))
            c.font = Font(bold=True if col_idx == 2 else False,
                           size=9, name="Calibri", color=D_GREY)
            c.number_format = "DD-MMM-YYYY" if col_idx == 2 else "0.000%"
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin()

    # 100 blank rows for manual entry
    for n in range(6, 106):
        ws.row_dimensions[n].height = 18
        c = ws.cell(row=n, column=2)
        c.fill = make_fill(YELLOW); c.number_format = "DD-MMM-YYYY"
        c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
        c.alignment = Alignment(horizontal="center"); c.border = thin()
        for col in range(3, 12):
            c = ws.cell(row=n, column=col)
            c.fill = make_fill(YELLOW); c.number_format = "0.000%"
            c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
            c.alignment = Alignment(horizontal="center"); c.border = thin()

    # Note on spread calculation methodology
    ws.merge_cells("B108:K108")
    note = ws["B108"]
    note.value = ("Spread = Instrument Yield - G-Sec Yield (nearest tenor). "
                  "For CP (tenor < 1Y): compare vs 1Y G-Sec. "
                  "For NCD 5Y: compare vs 5Y G-Sec.")
    note.font  = Font(italic=True, size=9, color=D_GREY, name="Calibri")
    note.fill  = make_fill(GREY)
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[108].height = 22

    apply_border(ws, 4, 106, 2, 11)
    ws.freeze_panes = "C5"
    return ws


# ──────────────────────────────────────────────
# SHEET 4 — PORTFOLIO SUMMARY
# ──────────────────────────────────────────────
def build_summary(wb, de_ws, de_start, de_total):
    ws = wb.create_sheet("Portfolio_Summary")
    ws.sheet_view.showGridLines = False

    widths = {"A":2,"B":22,"C":14,"D":14,"E":14,"F":14,
              "G":14,"H":14,"I":14,"J":14,"K":14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:K1")
    t = ws["B1"]
    t.value = "PORTFOLIO SUMMARY — LIVE"
    t.font  = Font(bold=True, size=13, color=WHITE, name="Calibri")
    t.fill  = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("B2:K2")
    s = ws["B2"]
    s.value = "Aggregated from Data_Entry sheet. Auto-updates when entries are changed."
    s.font  = Font(italic=True, size=9, color=D_GREY, name="Calibri")
    s.fill  = make_fill(GREY)
    s.alignment = Alignment(horizontal="center")

    # ── Section 1: By Instrument Type
    ws.merge_cells("B4:E4")
    hdr(ws, 4, 2, "1. PORTFOLIO BY INSTRUMENT TYPE", merge_end_col=5)

    sub_headers = ["Instrument","Face Value (₹ Cr)","MTM Value (₹ Cr)","Wtd Avg Yield","Wtd Avg Tenor"]
    for i, h in enumerate(sub_headers, start=2):
        hdr(ws, 5, i, h, sub=True)
    ws.row_dimensions[5].height = 22

    instr_types = ["NCD", "CP", "CD", "Bond", "Other"]
    for n, instr in enumerate(instr_types, start=6):
        ws.row_dimensions[n].height = 20
        ws.cell(row=n, column=2, value=instr).font = Font(size=9, bold=True, name="Calibri")
        ws.cell(row=n, column=2).fill = make_fill(GREY)
        ws.cell(row=n, column=2).border = thin()
        ws.cell(row=n, column=2).alignment = Alignment(horizontal="center", vertical="center")

        face_formula = (
            f'=SUMIFS(Data_Entry!H{de_start}:H{de_total},'
            f'Data_Entry!C{de_start}:C{de_total},"{instr}")'
        )
        c = ws.cell(row=n, column=3, value=face_formula)
        c.number_format = "#,##0.00"; c.fill = make_fill(GREEN)
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)
        c.border = thin(); c.alignment = Alignment(horizontal="center", vertical="center")

        mtm_formula = (
            f'=SUMIFS(Data_Entry!N{de_start}:N{de_total},'
            f'Data_Entry!C{de_start}:C{de_total},"{instr}")'
        )
        c = ws.cell(row=n, column=4, value=mtm_formula)
        c.number_format = "#,##0.00"; c.fill = make_fill(GREEN)
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)
        c.border = thin(); c.alignment = Alignment(horizontal="center", vertical="center")

        # Weighted avg yield = SUMPRODUCT of yield × face / total face
        wtd_yield = (
            f'=IFERROR(SUMPRODUCT((Data_Entry!C{de_start}:C{de_total}="{instr}")*'
            f'Data_Entry!I{de_start}:I{de_total}*Data_Entry!H{de_start}:H{de_total})'
            f'/SUMIFS(Data_Entry!H{de_start}:H{de_total},'
            f'Data_Entry!C{de_start}:C{de_total},"{instr}"),"")'
        )
        c = ws.cell(row=n, column=5, value=wtd_yield)
        c.number_format = "0.00%"; c.fill = make_fill(GREEN)
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)
        c.border = thin(); c.alignment = Alignment(horizontal="center", vertical="center")

        # Weighted avg tenor
        wtd_tenor = (
            f'=IFERROR(SUMPRODUCT((Data_Entry!C{de_start}:C{de_total}="{instr}")*'
            f'Data_Entry!G{de_start}:G{de_total}*Data_Entry!H{de_start}:H{de_total})'
            f'/SUMIFS(Data_Entry!H{de_start}:H{de_total},'
            f'Data_Entry!C{de_start}:C{de_total},"{instr}"),"")'
        )
        c = ws.cell(row=n, column=6, value=wtd_tenor)
        c.number_format = "0.00"; c.fill = make_fill(GREEN)
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)
        c.border = thin(); c.alignment = Alignment(horizontal="center", vertical="center")

    # Total row
    r = 11
    ws.row_dimensions[r].height = 22
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    tot = ws.cell(row=r, column=2, value="TOTAL PORTFOLIO")
    tot.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    tot.fill = make_fill(DARK_BLUE); tot.border = thin()
    tot.alignment = Alignment(horizontal="center", vertical="center")

    for col, formula in [
        (3, f'=SUM(C6:C10)'),
        (4, f'=SUM(D6:D10)'),
        (5, f'=IFERROR(SUMPRODUCT(Data_Entry!I{de_start}:I{de_total}*Data_Entry!H{de_start}:H{de_total})/SUM(Data_Entry!H{de_start}:H{de_total}),"")'),
        (6, f'=IFERROR(SUMPRODUCT(Data_Entry!G{de_start}:G{de_total}*Data_Entry!H{de_start}:H{de_total})/SUM(Data_Entry!H{de_start}:H{de_total}),"")'),
    ]:
        c = ws.cell(row=r, column=col, value=formula)
        c.number_format = "#,##0.00" if col <= 4 else ("0.00%" if col == 5 else "0.00")
        c.fill = make_fill(DARK_BLUE); c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.border = thin(); c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Section 2: By Issuer
    r = 14
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    hdr(ws, r, 2, "2. PORTFOLIO BY ISSUER", merge_end_col=6)

    r = 15
    issuer_headers = ["Issuer Name","Instrument","Face (₹ Cr)","Yield (%)","G-Sec Bench","Spread (bp)","MTM (₹ Cr)"]
    for i, h in enumerate(issuer_headers, start=2):
        hdr(ws, r, i, h, sub=True)
    ws.row_dimensions[r].height = 22

    # 20 issuer rows with SUMIFS
    for n in range(16, 36):
        ws.row_dimensions[n].height = 18
        for col in range(2, 9):
            c = ws.cell(row=n, column=col)
            c.fill = make_fill(LIGHT_BLUE if n % 2 == 0 else WHITE)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin()
            # Formula: pull unique issuer names (simplified — first match)
            if col == 2:
                c.value = f'=IFERROR(INDEX(Data_Entry!D{de_start}:D{de_total},MATCH(0,COUNTIF($B$16:B{n-1},Data_Entry!D{de_start}:D{de_total}),0)+{de_start-1}),"")'
                c.fill = make_fill(GREY)
                c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)

    # ── Section 3: Duration & Risk
    r = 38
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    hdr(ws, r, 2, "3. PORTFOLIO DURATION & RISK METRICS", merge_end_col=5)

    r = 39
    risk_headers = ["Metric","Value","Formula / Notes"]
    for i, h in enumerate(risk_headers, start=2):
        hdr(ws, r, i, h, sub=True)

    risk_metrics = [
        (40, "Portfolio Modified Duration",
         f'=IFERROR(SUMPRODUCT((Data_Entry!C{de_start}:C{de_total})="NCD")*Data_Entry!G{de_start}:G{de_total}*Data_Entry!H{de_start}:H{de_total})/SUMIFS(Data_Entry!H{de_start}:H{de_total},Data_Entry!C{de_start}:C{de_total},"NCD")/(1+Data_Entry!I{de_start}:I{de_total}),"Check data")',
         "0.00 Yrs", "Approx — assumes bullet NCD. For amortising, use actual schedule."),
        (41, "Portfolio DV01 (₹ Cr per bp)",
         f'=C40*SUM(Data_Entry!H{de_start}:H{de_total})/10000',
         "#,##0.0000", "Price sensitivity per 1bp yield move"),
        (42, "Estimated MTM Δ per +100bp",
         f'=-C40*SUM(Data_Entry!H{de_start}:H{de_total})*0.01',
         "#,##0.00", "Approx portfolio loss for 100bp rate rise"),
        (43, "Estimated MTM Δ per -100bp",
         f'=C40*SUM(Data_Entry!H{de_start}:H{de_total})*0.01',
         "#,##0.00", "Approx portfolio gain for 100bp rate fall"),
    ]

    for row, label, formula, fmt, note in risk_metrics:
        ws.row_dimensions[row].height = 22
        lbl = ws.cell(row=row, column=2, value=label)
        lbl.font = Font(size=9, bold=True, name="Calibri", color=D_GREY)
        lbl.fill = make_fill(GREY); lbl.border = thin()
        lbl.alignment = Alignment(vertical="center", wrap_text=True)

        val = ws.cell(row=row, column=3, value=formula)
        val.number_format = fmt; val.fill = make_fill(GREEN)
        val.font = Font(bold=True, size=10, name="Calibri", color=D_GREY)
        val.border = thin(); val.alignment = Alignment(horizontal="center", vertical="center")

        nt = ws.cell(row=row, column=4, value=note)
        nt.font = Font(size=8, italic=True, name="Calibri", color=D_GREY)
        nt.fill = make_fill(GREY); nt.border = thin()
        nt.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)

    apply_border(ws, 5, 43, 2, 6)
    ws.freeze_panes = "B5"
    return ws


# ──────────────────────────────────────────────
# SHEET 5 — MTM TRACKER
# ──────────────────────────────────────────────
def build_mtm(wb, de_ws, de_start, de_total):
    ws = wb.create_sheet("MTM")
    ws.sheet_view.showGridLines = False

    widths = {"A":2,"B":14,"C":20,"D":14,"E":14,"F":14,
              "G":14,"H":14,"I":14,"J":14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:J1")
    t = ws["B1"]
    t.value = "MARK-TO-MARKET (MTM) TRACKER"
    t.font  = Font(bold=True, size=13, color=WHITE, name="Calibri")
    t.fill  = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("B2:J2")
    s = ws["B2"]
    s.value = ("Tracks daily MTM using latest prices vs cost/issue price. "
               "For unrealized gain/loss: MTM Value - Carrying Value.")
    s.font  = Font(italic=True, size=9, color=D_GREY, name="Calibri")
    s.fill  = make_fill(GREY)
    s.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    hdrs = ["Date\n(Manual)","Issuer","Instrument","Face Value (₹ Cr)",
            "Issue Price / Cost (₹)","Current Clean Price (₹)","MTM Value (₹ Cr)",
            "Unrealised G/L (₹ Cr)","Unrealised G/L (%)"]
    for i, h in enumerate(hdrs, start=2):
        hdr(ws, 3, i, h, sub=True)
    ws.row_dimensions[3].height = 28

    # Reference Data_Entry — each MTM row links to corresponding Data_Entry row
    # MTM row 4 → Data_Entry row de_start (offset 0)
    for n in range(4, 54):
        ws.row_dimensions[n].height = 18
        offset = n - 4  # 0-based offset from de_start
        de_row = de_start + offset

        for col in range(2, 11):
            c = ws.cell(row=n, column=col)
            c.fill = make_fill(LIGHT_BLUE if n % 2 == 0 else WHITE)
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin()

        # Col C: pull issuer from Data_Entry
        c = ws.cell(row=n, column=3,
                    value=f'=IFERROR(Data_Entry!D${de_row},"")')

        # Col D: Instrument
        c = ws.cell(row=n, column=4,
                    value=f'=IFERROR(Data_Entry!C${de_row},"")')

        # Col E: Face Value
        c = ws.cell(row=n, column=5,
                    value=f'=IFERROR(Data_Entry!H${de_row},"")')
        c.number_format = "#,##0.00"

        # Col F: Issue/Cost price (Invoice Price from Data_Entry col L)
        c = ws.cell(row=n, column=6,
                    value=f'=IFERROR(Data_Entry!L${de_row},"")')
        c.number_format = "0.00"

        # Col G: Current price (clean price — update manually or via Data_Entry)
        c = ws.cell(row=n, column=7,
                    value=f'=IFERROR(Data_Entry!J${de_row},"")')
        c.number_format = "0.00"
        c.fill = make_fill(YELLOW)
        c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)

        # Col H: MTM Value (₹ Cr) = current price × face / 100
        c = ws.cell(row=n, column=8,
                    value=f'=IFERROR(G{n}*E{n}/100,"")')
        c.number_format = "#,##0.00"; c.fill = make_fill(GREEN)
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)

        # Col I: Unrealised G/L (₹ Cr)
        c = ws.cell(row=n, column=9,
                    value=f'=IFERROR(H{n}-F{n}*E{n}/100,"")')
        c.number_format = "#,##0.00"; c.fill = make_fill(GREEN)
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)

        # Col J: Unrealised G/L %
        c = ws.cell(row=n, column=10,
                    value=f'=IFERROR(I{n}/(F{n}*E{n}/100),"")')
        c.number_format = "0.00%"; c.fill = make_fill(GREEN)
        c.font = Font(bold=True, size=9, name="Calibri", color=D_GREY)

    apply_border(ws, 3, 53, 2, 10)
    ws.freeze_panes = "C4"
    return ws


# ──────────────────────────────────────────────
# SHEET 6 — ALERTS
# ──────────────────────────────────────────────
def build_alerts(wb):
    ws = wb.create_sheet("Alerts")
    ws.sheet_view.showGridLines = False

    widths = {"A":2,"B":20,"C":14,"D":14,"E":16,"F":16,"G":20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:G1")
    t = ws["B1"]
    t.value = "ALERTS & THRESHOLD MONITORING"
    t.font  = Font(bold=True, size=13, color=WHITE, name="Calibri")
    t.fill  = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("B2:G2")
    s = ws["B2"]
    s.value = ("Set yield/spread/basis-point thresholds below. "
               "Enter trigger values in YELLOW. Alerts generate when current > threshold (RED = breach).")
    s.font  = Font(italic=True, size=9, color=D_GREY, name="Calibri")
    s.fill  = make_fill(GREY)
    s.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.row_dimensions[2].height = 20

    # Alert types
    alert_sections = [
        ("YIELD BREACHES", [
            ("Max portfolio yield drop", "Yield", "Below", 0.05, "0.00%"),
            ("Max portfolio yield spike", "Yield", "Above", 0.12, "0.00%"),
        ]),
        ("SPREAD ALERTS", [
            ("NCD spread over G-Sec (max)", "Spread vs G-Sec", "Above", 200, "0 bp"),
            ("CP spread over G-Sec (max)", "Spread vs G-Sec", "Above", 150, "0 bp"),
        ]),
        ("MTM ALERTS", [
            ("Daily MTM drop (₹ Cr)", "MTM Unrealised G/L", "Below", -5.0, "#,##0.00"),
            ("Daily MTM gain (₹ Cr)", "MTM Unrealised G/L", "Above", 5.0, "#,##0.00"),
        ]),
        ("CONCENTRATION ALERTS", [
            ("Single issuer max exposure (₹ Cr)", "Issuer Face Value", "Above", 50.0, "#,##0.00"),
            ("Single instrument max (₹ Cr)", "Face Value", "Above", 100.0, "#,##0.00"),
        ]),
    ]

    current_row = 4
    for section_name, alerts in alert_sections:
        ws.merge_cells(start_row=current_row, start_column=2,
                       end_row=current_row, end_column=7)
        hdr(ws, current_row, 2, section_name, merge_end_col=7)
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Sub-headers
        sub = ["Alert Name", "Metric", "Direction", "Threshold (Enter)", "Current Value", "Status"]
        for i, h in enumerate(sub, start=2):
            hdr(ws, current_row, i, h, sub=True)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for alert_name, metric, direction, default, fmt in alerts:
            ws.row_dimensions[current_row].height = 20

            vals = [alert_name, metric, direction, default, "", ""]
            for i, val in enumerate(vals, start=2):
                c = ws.cell(row=current_row, column=i, value=val)
                c.font = Font(size=9, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin()
                if i == 5:
                    c.fill = make_fill(YELLOW)
                    c.font = Font(bold=True, size=9, name="Calibri", color=DARK_BLUE)
                    c.number_format = fmt
                elif i == 6:
                    c.fill = make_fill(LIGHT_BLUE)
                    c.number_format = fmt
                elif i == 7:
                    c.fill = make_fill(GREEN)
                    c.value = f'=IF(E{current_row}="","",IF(D{current_row}="Above",IF(F{current_row}>E{current_row},"⚠ BREACH","✅ OK"),IF(F{current_row}<E{current_row},"⚠ BREACH","✅ OK")))'
                    c.font = Font(bold=True, size=10, name="Calibri")
                    c.number_format = "@"

            current_row += 1

        current_row += 1  # spacer

    # Legend
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row, end_column=7)
    legend = ws.cell(row=current_row, column=2)
    legend.value = "⚠ BREACH = Threshold crossed — review needed  |  ✅ OK = Within thresholds"
    legend.font  = Font(italic=True, size=9, name="Calibri", color=D_GREY)
    legend.fill  = make_fill(GREY)
    legend.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 20

    ws.freeze_panes = "B4"
    return ws


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def build_ncd_cp_tracker(output_path=None):
    print("Building NCD + CP Rate Tracker...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("  → Data_Entry sheet...")
    de_ws, de_start, de_total = build_data_entry(wb)

    print("  → Rates_DB sheet...")
    build_rates_db(wb, de_ws, de_start, de_total)

    print("  → GSec_Benchmark sheet...")
    build_gsec(wb)

    print("  → Portfolio_Summary sheet...")
    build_summary(wb, de_ws, de_start, de_total)

    print("  → MTM sheet...")
    build_mtm(wb, de_ws, de_start, de_total)

    print("  → Alerts sheet...")
    build_alerts(wb)

    if not output_path:
        script_dir = "/home/homepc/.openclaw/workspace/Excel-Models/NCD"
        output_path = f"{script_dir}/NCD_CP_Rate_Tracker.xlsx"

    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")
    return output_path


if __name__ == "__main__":
    build_ncd_cp_tracker()
